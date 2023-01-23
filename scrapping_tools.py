import json
import urllib3

import requests
import lxml

from config import *
from condition import check_scores
from time import sleep
from openpyxl.styles import PatternFill, Alignment
from bs4 import BeautifulSoup, element
from fake_useragent import UserAgent
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.remote.webelement import WebElement



class Basketball:
    """Class contains tools for scrapping"""

    def __init__(self):
        self.options = webdriver.ChromeOptions()
        s = Service(executable_path="chromedriver.exe")
        self.ua = UserAgent(browsers=["chrome"])
        self.wb = WB
        self.sheet = SHEET

        self.options.add_argument(
            f'user-agent={self.ua.random}'
        )
        self.options.add_argument(
            '--disable-blink-features=AutomationControlled'
        )
        self.options.add_argument(
            '--ignore-certificate-errors-spki-list'
        )
        self.options.add_argument('--ignore-ssl-errors')
        self.options.add_argument('--headless')
        self.options.add_argument('log-level=2')
        self.driver = webdriver.Chrome(
            service=s, options=self.options
        )
        self.window_before = self._log_in_to_site()


    def _log_in_to_site(self) -> str:
        # Authorization on the website
        # Авторизация на сайте
        self.driver.get(url=START_PAGE)
        window = self.driver.window_handles[0]
        sleep(3)
        try:
            self.driver.find_element(By.ID, 'onetrust-accept-btn-handler').click()
        except NoSuchElementException:
            pass

        self.__log_in_account()
        return window
 

    def __log_in_account(self):
        # Account authorization
        # Авторизация в аккаунте
        self.driver.find_element(By.ID, 'header__block--user-menu').click()
        sleep(1.5)

        try:
            self.driver.find_element(By.XPATH, '//*[@id="header__block--user-menu"]/div[2]/div/div/div/section[2]/button[4]').click()
        except NoSuchElementException:
            pass

        input_email = self.driver.find_element(By.XPATH, '//*[@id="email"]')
        input_email.clear()
        input_email.send_keys(LOGIN)

        input_password = self.driver.find_element(By.XPATH, '//*[@id="passwd"]')
        input_password.clear()
        input_password.send_keys(PASSWORD)

        WebDriverWait(self.driver, 5).until(
            EC.element_to_be_clickable(
                (
                    By.XPATH,
                    '//*[@id="header__block--user-menu"]/div[2]/div/div/div/section/button'
                )
            )
        ).click()
        sleep(7)


    def _get_score_1_q(self, row: int, scores_1t: element.ResultSet,
                       scores_2t: element.ResultSet) -> None:
        # Get the score at the first quarter and add up the total points
        # Получить счет при первой четверти и сложить общее количество очков
        self.sheet[f'E{row}'] = int(scores_1t[0].text.strip())
        self.sheet[f'E{row + 1}'] = int(scores_2t[0].text.strip())
        self.sheet[f'I{row}'] = int(self.sheet[f'E{row}'].value)
        self.sheet[f'I{row + 1}'] = int(self.sheet[f'E{row + 1}'].value)

    def _get_score_2_q(self, row: int, scores_1t: element.ResultSet,
                       scores_2t: element.ResultSet) -> None:
        # Get the score at the second quarter and add up the total points
        # Получить счет при второй четверти и сложить общее количество очков
        self.sheet[f'F{row}'] = int(scores_1t[1].text.strip())
        self.sheet[f'F{row + 1}'] = int(scores_2t[1].text.strip())
        self.sheet[f'I{row}'] = int(self.sheet[f'E{row}'].value) \
            + int(self.sheet[f'F{row}'].value)
        self.sheet[f'I{row + 1}'] = int(self.sheet[f'E{row + 1}'].value) \
            + int(self.sheet[f'F{row + 1}'].value)

    def _get_score_3_q(self, row: int, scores_1t: element.ResultSet,
                       scores_2t: element.ResultSet) -> None:
        # Get the score at the third quarter and add up the total points
        # Получить счет при третьей четверти и сложить общее количество очков
        self.sheet[f'G{row}'] = int(scores_1t[2].text.strip())
        self.sheet[f'G{row + 1}'] = int(scores_2t[2].text.strip())
        self.sheet[f'I{row}'] = int(self.sheet[f'E{row}'].value) \
            + int(self.sheet[f'F{row}'].value) \
            + int(self.sheet[f'G{row}'].value)
        self.sheet[f'I{row + 1}'].value = int(self.sheet[f'E{row + 1}'].value) \
            + int(self.sheet[f'F{row + 1}'].value) \
            + int(self.sheet[f'G{row + 1}'].value)

    def _get_score_4_q(self, row: int, scores_1t: element.ResultSet,
                       scores_2t: element.ResultSet) -> None:
        # Get the score at the fourth quarter and add up the total points
        # Получить счет при четвертой четверти и сложить общее количество очков
        self.sheet[f'H{row}'] = int(scores_1t[3].text.strip())
        self.sheet[f'H{row + 1}'] = int(scores_2t[3].text.strip())
        self.sheet[f'I{row}'] = int(self.sheet[f'E{row}'].value) \
            + int(self.sheet[f'F{row}'].value) \
            + int(self.sheet[f'G{row}'].value) \
            + int(self.sheet[f'H{row}'].value)
        self.sheet[f'I{row + 1}'].value = int(self.sheet[f'E{row + 1}'].value) \
            + int(self.sheet[f'F{row + 1}'].value) \
            + int(self.sheet[f'G{row + 1}'].value) \
            + int(self.sheet[f'H{row + 1}'].value)


    def _get_coeffs(self, row: int,
                    coefficients: element.ResultSet) -> None:
        # Get the coefficients and record the difference
        # Получить коэффициенты и записать разницу
        cf_1 = float(
            coefficients[0].find('span', class_='oddsValueInner').text.strip()
        )
        cf_2 = float(
            coefficients[1].find('span', class_='oddsValueInner').text.strip()
        )
        with open('coeffs_transl_tb.json', 'r', encoding='utf-8') as file:
            cf_tr = json.load(file)
            if cf_1 < cf_2:
                self.sheet[f'C{row}'] = cf_tr.get(str(cf_1))
                self.sheet[f'C{row}'].alignment = Alignment(
                    horizontal='center', vertical='center'
                )
            elif cf_1 > cf_2:
                self.sheet[f'C{row + 1}']= cf_tr.get(str(cf_2))
                self.sheet[f'C{row + 1}'].alignment = Alignment(
                    horizontal='center', vertical='center'
                )


    def _get_game_number(self, row: int) -> None:
        # Calculate and record the game number in the table
        # Рассчитать и записать номер матча в таблице
        previous_cell = self.sheet[f'A{row - 2}'].value
        if previous_cell:
            self.sheet.merge_cells(f'A{row}:A{row + 1}')
            self.sheet[f'A{row}'] = int(previous_cell) + 1
            self.sheet[f'A{row}'].alignment = Alignment(
                horizontal='center', vertical='center'
            )

    @staticmethod
    def send_message(message: str, chat_id: str, token: str) -> None:
        # Send a message in Telegram
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        url = f'https://api.telegram.org/bot{token}/sendMessage'
        requests.post(
            url=url,
            timeout=5,
            verify=False,
            data={
                'chat_id': chat_id,
                'text': message
            }
        )


    def check_message(self, row: int, time: str) -> None:
        # Check the difference and the sum
        # Проверить разницу и сумму
        send = False

        if (self.sheet[f'C{row}'].value is None) and (self.sheet[f'C{row + 1}'].value is None):
            pass
        else:
            if '1-я четверть 10' in time:
                if self.sheet[f'C{row}'].value is None:
                    df = float(self.sheet[f'C{row + 1}'].value)
                    sum = df + (int(self.sheet[f'E{row + 1}'].value) - int(self.sheet[f'E{row}'].value))
                    send = check_scores(
                        row=row + 1, sum=sum, column='E', df_1='R7', df_2='R9'
                    )
                else:
                    df = float(self.sheet[f'C{row}'].value)
                    sum = df + (int(self.sheet[f'E{row}'].value) - int(self.sheet[f'E{row + 1}'].value))
                    send = check_scores(
                        row=row, sum=sum, column='E', df_1='R7', df_2='R9'
                    )

            elif '2-я четверть 10' in time:
                if self.sheet[f'C{row}'].value is None:
                    df = float(self.sheet[f'C{row + 1}'].value)
                    sum = df + ((int(self.sheet[f'E{row + 1}'].value) + int(self.sheet[f'F{row + 1}'].value))
                        - (int(self.sheet[f'E{row}'].value) + int(self.sheet[f'F{row}'].value)))
                    send = check_scores(
                        row=row + 1, sum=sum, column='F', df_1='S7', df_2='S9'
                    )
                else:
                    df = float(self.sheet[f'C{row}'].value)
                    sum = df + ((int(self.sheet[f'E{row}'].value) + int(self.sheet[f'F{row}'].value))
                        - (int(self.sheet[f'E{row + 1}'].value) + int(self.sheet[f'F{row + 1}'].value)))
                    send = check_scores(
                        row=row, sum=sum, column='F', df_1='S7', df_2='S9'
                    )

            elif '3-я четверть 10' in time:
                if self.sheet[f'C{row}'].value is None:
                    df = float(self.sheet[f'C{row + 1}'].value)
                    sum = df + ((int(self.sheet[f'E{row + 1}'].value) + int(self.sheet[f'F{row + 1}'].value)
                        + int(self.sheet[f'G{row + 1}'].value)) - (int(self.sheet[f'E{row}'].value)
                        + int(self.sheet[f'F{row}'].value) + int(self.sheet[f'G{row}'].value)))
                    send = check_scores(
                        row=row + 1, sum=sum, column='G', df_1='T7', df_2='T9'
                    )
                else:
                    df = float(self.sheet[f'C{row}'].value)
                    sum = df + ((int(self.sheet[f'E{row}'].value) + int(self.sheet[f'F{row}'].value)
                        + int(self.sheet[f'G{row}'].value)) - (int(self.sheet[f'E{row + 1}'].value)
                        + int(self.sheet[f'F{row + 1}'].value) + int(self.sheet[f'G{row + 1}'].value)))
                    send = check_scores(
                        row=row, sum=sum, column='G', df_1='T7', df_2='T9'
                    )

            if send:
                team_1 = self.sheet[f'D{row}'].value
                score_1t = self.sheet[f'I{row}'].value
                team_2 = self.sheet[f'D{row + 1}'].value
                score_2t = self.sheet[f'I{row + 1}'].value
                table = self.sheet[f'J{row}'].value
                Basketball.send_message(
                    f'ВРЕМЯ СТАВОК🏀🏀🏀\nМАТЧ:\n{team_1} - {team_2}\n{score_1t} - {score_2t}\n'
                    f'{time}\nТУРНИРНАЯ ТАБЛИЦА: {table}',
                    CHAT_ID, TOKEN
                )


    def update_info(self, game: any, count_xl: int,
                    time: str, first_use: bool = False,
                    game_sl: WebElement = 0) -> None:
        # Update data for a one game
        # Обновить данные для одного матча
        if not first_use:
            self.sheet.merge_cells(f'B{count_xl}:B{count_xl + 1}')
        self.sheet[f'B{count_xl}'] = time
        self.sheet[f'B{count_xl}'].alignment = Alignment(
            horizontal='center', vertical='center'
        )

        scores_1t = game.find_all('div', class_='event__part--home')
        scores_2t = game.find_all('div', class_='event__part--away')

        if '1-я четверть' in time:
            self._get_score_1_q(count_xl, scores_1t, scores_2t)
            if (not first_use) and ('1-я четверть 10' in time):
                self.check_message(row=count_xl, time=time)
        elif ('2-я четверть' in time) or ('Перерыв' in time):
            self._get_score_1_q(count_xl, scores_1t, scores_2t)
            self._get_score_2_q(count_xl, scores_1t, scores_2t)
            if (not first_use) and ('2-я четверть 10' in time):
                self.check_message(row=count_xl, time=time)
        elif '3-я четверть' in time:
            self._get_score_1_q(count_xl, scores_1t, scores_2t)
            self._get_score_2_q(count_xl, scores_1t, scores_2t)
            self._get_score_3_q(count_xl, scores_1t, scores_2t)
            if (not first_use) and ('3-я четверть 10' in time):
                self.check_message(row=count_xl, time=time)
        elif ('4-я четверть' in time) or ('Завершен' in time):
            self._get_score_1_q(count_xl, scores_1t, scores_2t)
            self._get_score_2_q(count_xl, scores_1t, scores_2t)
            self._get_score_3_q(count_xl, scores_1t, scores_2t)
            self._get_score_4_q(count_xl, scores_1t, scores_2t)


        if first_use:
            self._get_game_number(count_xl)

            WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((game_sl))
            ).click()
            self.driver.switch_to.window(self.driver.window_handles[1])

            next_soup = BeautifulSoup(self.driver.page_source, 'lxml')
            coefficients = next_soup.find_all('div', class_='cellWrapper')
            if coefficients:
                self._get_coeffs(count_xl, coefficients)

            try:
                table_mesh = self.driver.find_element(By.XPATH,
                    "//*[@id='detail']/div[6]/div/a[last()]")
            except NoSuchElementException:
                table_mesh = self.driver.find_element(By.XPATH,
                    "//*[@id='detail']/div[7]/div/a[last()]")

            WebDriverWait(self.driver, 8).until(
                EC.element_to_be_clickable((table_mesh))
            ).click()

            self.sheet.merge_cells(f'J{count_xl}:J{count_xl + 1}')
            self.sheet[f'J{count_xl}'] = self.driver.current_url

            self.driver.close()
            self.driver.switch_to.window(self.window_before)

        self.wb.save(EXCEL_FILE)


    def _close_driver(self):
        self.driver.close()
        self.driver.quit()


